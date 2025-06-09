from openpyxl import load_workbook
from twilio.rest import Client
import config

# config file is available on local machine, not uploaded to GitHub for security reasons
client = Client(config.TWILIO_ACCOUNT_SID, config.TWILIO_AUTH_TOKEN)

wb = load_workbook('sample.xlsx')
sheet = wb.active

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    
    if name:
        message = f"Dear parent, this is a demo invoice for {name}. Due date: June 8, 2025. Thank you!"
        client.messages.create(
            body=message,
            from_=config.TWILIO_PHONE_NUMBER,
            to="+16784885297" 
        )
        print(f"Sent message to {name}")